import io
import openpyxl
from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from .models import Parto, RN

# === REPORTE 1: Características del Parto ===

def exportar_reporte_parto_pdf(request):
    print("grafico_base64:", bool(request.POST.get("grafico_base64")))

    data = {
        "headers": ["Tipo de Parto", "Cantidad"],
        "rows": [
            ["Total Partos", Parto.objects.count()],
            ["Vaginal", Parto.objects.filter(tipo_parto="vaginal").count()],
            ["Instrumental", Parto.objects.filter(tipo_parto="instrumental").count()],
            ["Cesárea Electiva", Parto.objects.filter(tipo_parto="cesarea_electiva").count()],
            ["Cesárea Urgencia", Parto.objects.filter(tipo_parto="cesarea_urgencia").count()],
            ["Lactancia precoz (≥2500 g)", RN.objects.filter(peso__gte=2500, lactancia_antes_60=True).count()],
        ],
        "grafico_base64": request.POST.get("grafico_base64")
    }

    html = render_to_string("reportes/exportables/rem_24_caracteristicas_parto_pdf.html", {"data": data})
    result = io.BytesIO()
    pisa.CreatePDF(html, dest=result)

    response = HttpResponse(result.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_parto.pdf"'
    return response


def exportar_reporte_parto_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Características del Parto"

    headers = ["Tipo de Parto", "Cantidad"]
    rows = [
        ["Total Partos", Parto.objects.count()],
        ["Vaginal", Parto.objects.filter(tipo_parto="vaginal").count()],
        ["Instrumental", Parto.objects.filter(tipo_parto="instrumental").count()],
        ["Cesárea Electiva", Parto.objects.filter(tipo_parto="cesarea_electiva").count()],
        ["Cesárea Urgencia", Parto.objects.filter(tipo_parto="cesarea_urgencia").count()],
        ["Lactancia precoz (≥2500 g)", RN.objects.filter(peso__gte=2500, lactancia_antes_60=True).count()],
    ]

    # Estilos
    from openpyxl.styles import Font, Alignment, PatternFill
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    center_align = Alignment(horizontal="center")

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for row in rows:
        ws.append(row)

    # Ajuste de ancho
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="reporte_parto.xlsx"'
    wb.save(response)
    return response


# === REPORTE 2: Información general RN vivos ===

def exportar_reporte_nacidos_vivos_pdf(request):
    data = {
        "headers": ["Rango de Peso (g)", "Cantidad"],
        "rows": [
            ["Menos de 500", RN.objects.filter(peso__lt=500).count()],
            ["500 a 999", RN.objects.filter(peso__gte=500, peso__lte=999).count()],
            ["1.000 a 1.499", RN.objects.filter(peso__gte=1000, peso__lte=1499).count()],
            ["1.500 a 1.999", RN.objects.filter(peso__gte=1500, peso__lte=1999).count()],
            ["2.000 a 2.499", RN.objects.filter(peso__gte=2000, peso__lte=2499).count()],
            ["2.500 a 2.999", RN.objects.filter(peso__gte=2500, peso__lte=2999).count()],
            ["3.000 a 3.999", RN.objects.filter(peso__gte=3000, peso__lte=3999).count()],
            ["4.000 y más", RN.objects.filter(peso__gte=4000).count()],
            ["Total nacidos vivos", RN.objects.count()],
            ["Con anomalía congénita", RN.objects.filter(anomalia_congenita=True).count()],
        ],
        "grafico_base64": request.POST.get("grafico_base64")
    }

    html = render_to_string("reportes/exportables/rem_a24_info_gral_rn_vivos_pdf.html", {"data": data})
    result = io.BytesIO()
    pisa.CreatePDF(html, dest=result)

    response = HttpResponse(result.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_nacidos_vivos.pdf"'
    return response


def exportar_reporte_nacidos_vivos_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Info RN Vivos"

    headers = ["Rango de Peso (g)", "Cantidad"]
    rows = [
        ["Menos de 500", RN.objects.filter(peso__lt=500).count()],
        ["500 a 999", RN.objects.filter(peso__gte=500, peso__lte=999).count()],
        ["1.000 a 1.499", RN.objects.filter(peso__gte=1000, peso__lte=1499).count()],
        ["1.500 a 1.999", RN.objects.filter(peso__gte=1500, peso__lte=1999).count()],
        ["2.000 a 2.499", RN.objects.filter(peso__gte=2000, peso__lte=2499).count()],
        ["2.500 a 2.999", RN.objects.filter(peso__gte=2500, peso__lte=2999).count()],
        ["3.000 a 3.999", RN.objects.filter(peso__gte=3000, peso__lte=3999).count()],
        ["4.000 y más", RN.objects.filter(peso__gte=4000).count()],
        ["Total nacidos vivos", RN.objects.count()],
        ["Con anomalía congénita", RN.objects.filter(anomalia_congenita=True).count()],
    ]

    from openpyxl.styles import Font, Alignment, PatternFill
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    center_align = Alignment(horizontal="center")

    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for row in rows:
        ws.append(row)

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="reporte_nacidos_vivos.xlsx"'
    wb.save(response)
    return response


# === REPORTE 3: Atención inmediata RN ===

def exportar_reporte_atencion_inmediata_pdf(request):
    data = {
        "headers": ["Indicador", "Cantidad"],
        "rows": [
            ["Total nacidos vivos", RN.objects.count()],
            ["Profilaxis Hepatitis B", RN.objects.filter(vacuna_hepatitis_b=True).count()],
            ["Profilaxis Ocular", RN.objects.filter(profilaxis_ocular=True).count()],
            ["Parto Vaginal", Parto.objects.filter(tipo_parto="vaginal").count()],
            ["Parto Instrumental", Parto.objects.filter(tipo_parto="instrumental").count()],
            ["Cesárea", Parto.objects.filter(tipo_parto__startswith="cesarea").count()],
            ["Parto Extrahospitalario", Parto.objects.filter(tipo_parto="extrahospitalario").count()],
            ["Apgar ≤ 3 al minuto", RN.objects.filter(apgar_1__lte=3).count()],
            ["Apgar ≤ 6 a los 5 minutos", RN.objects.filter(apgar_5__lte=6).count()],
            ["Reanimación Básica", RN.objects.filter(reanimacion_basica=True).count()],
            ["Reanimación Avanzada", RN.objects.filter(reanimacion_avanzada=True).count()],
            ["EHI Grado II y III", RN.objects.filter(ehi_grado_ii_iii=True).count()],
        ],
        "grafico_base64": request.POST.get("grafico_base64")
    }

    html = render_to_string("reportes/exportables/rem_a24_atencion_inmediata_rn_pdf.html", {"data": data})
    result = io.BytesIO()
    pisa.CreatePDF(html, dest=result)

    response = HttpResponse(result.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_atencion_inmediata.pdf"'
    return response


def exportar_reporte_atencion_inmediata_excel(request):
    from django.db.models.functions import TruncMonth
    from django.db.models import Count
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Atención Inmediata RN"
    ws.page_setup.orientation = 'landscape'

    # Encabezados horizontales
    headers = [
        "Mes",
        "Total nacidos vivos",
        "Profilaxis Hepatitis B",
        "Profilaxis Ocular",
        "Parto Vaginal",
        "Parto Instrumental",
        "Cesárea",
        "Parto Extrahospitalario",
        "Apgar ≤ 3 al minuto",
        "Apgar ≤ 6 a los 5 minutos",
        "Reanimación Básica",
        "Reanimación Avanzada",
        "EHI Grado II y III"
    ]
    ws.append(headers)

    # Estilos de encabezado
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    center_align = Alignment(horizontal="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Agrupar por mes
    meses = RN.objects.annotate(mes=TruncMonth('fecha_nacimiento')).values('mes').distinct().order_by('mes')

    for mes_obj in meses:
        mes = mes_obj['mes']
        rn_mes = RN.objects.filter(fecha_nacimiento__month=mes.month, fecha_nacimiento__year=mes.year)
        parto_mes = Parto.objects.filter(fecha_parto__month=mes.month, fecha_parto__year=mes.year)

        fila = [
            mes.strftime("%B %Y"),
            rn_mes.count(),
            rn_mes.filter(vacuna_hepatitis_b=True).count(),
            rn_mes.filter(profilaxis_ocular=True).count(),
            parto_mes.filter(tipo_parto="vaginal").count(),
            parto_mes.filter(tipo_parto="instrumental").count(),
            parto_mes.filter(tipo_parto__startswith="cesarea").count(),
            parto_mes.filter(tipo_parto="extrahospitalario").count(),
            rn_mes.filter(apgar_1__lte=3).count(),
            rn_mes.filter(apgar_5__lte=6).count(),
            rn_mes.filter(reanimacion_basica=True).count(),
            rn_mes.filter(reanimacion_avanzada=True).count(),
            rn_mes.filter(ehi_grado_ii_iii=True).count(),
        ]
        ws.append(fila)

    # Ajuste de ancho
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename=\"reporte_atencion_inmediata_horizontal.xlsx\"'
    wb.save(response)
    return response