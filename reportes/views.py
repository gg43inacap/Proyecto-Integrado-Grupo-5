# reportes/views.py
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .utils import get_reporte_rem24_completo
from django.template.loader import render_to_string
import io
# from xhtml2pdf import pisa  # Comentado temporalmente
from django.shortcuts import render
from datetime import datetime, date
from xhtml2pdf import pisa


# --- Nombres de meses en español ---
MESES_NOMBRES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

def nombre_mes(mes_num):
    return MESES_NOMBRES.get(mes_num, "Mes no especificado")

# --- Borde fino institucional ---
borde_fino = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# --- Helper para secciones horizontales ---
def write_horizontal_section(ws, start_row, titulo, rows):
    total_cols = len(rows)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=total_cols)
    cell_titulo = ws.cell(row=start_row, column=1, value=titulo)
    cell_titulo.font = Font(bold=True, size=12)
    cell_titulo.alignment = Alignment(horizontal="center")
    cell_titulo.border = borde_fino
    start_row += 1

    # Encabezados
    for idx, (label, _) in enumerate(rows, start=1):
        cell = ws.cell(row=start_row, column=idx, value=label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, horizontal="center")
        cell.border = borde_fino
    start_row += 1

    # Valores
    for idx, (_, value) in enumerate(rows, start=1):
        cell = ws.cell(row=start_row, column=idx, value=value)
        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = borde_fino
    start_row += 2

    return start_row

# --- Exportar Excel ---
def exportar_rem_a24_excel(request):
    # --- Capturar filtros desde GET ---
    mes_str = request.GET.get("mes")
    anio_str = request.GET.get("anio")
    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")

    mes = int(mes_str) if mes_str and mes_str.isdigit() else datetime.now().month
    anio = int(anio_str) if anio_str and anio_str.isdigit() else datetime.now().year
    inicio = date.fromisoformat(inicio) if inicio else None
    fin = date.fromisoformat(fin) if fin else None

    # --- Obtener datos del reporte ---
    data = get_reporte_rem24_completo(mes=mes, anio=anio, inicio=inicio, fin=fin)

    # --- Crear Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "REM A24"

    # Freeze panes
    ws.freeze_panes = "A4"

    # Ajuste de ancho
    for col in range(1, 20):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18

    # Título principal
    titulo_mes = nombre_mes(mes)


# Título principal con fondo suave
    ws.merge_cells("A1:E1")
    ws["A1"] = "REM A24 — Atención del Recién Nacido"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    ws.merge_cells("A2:E2")
    ws["A2"] = f"Reporte mensual — {titulo_mes} {anio}"
    ws["A2"].font = Font(bold=True, size=12, italic=True, color="404040")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    row = 4

    # --- Encabezado institucional tipo y total ---
    ws.cell(row=row, column=1, value="Tipo de reporte")
    ws.cell(row=row, column=2, value="REM A24 — Atención del Recién Nacido")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="left")
    ws.cell(row=row, column=1).border = borde_fino
    ws.cell(row=row, column=2).border = borde_fino
    row += 1

    ws.cell(row=row, column=1, value="Total nacidos vivos")
    ws.cell(row=row, column=2, value=data["seccion_d1"]["rows"][-2][1])  # penúltimo valor
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=1).border = borde_fino
    ws.cell(row=row, column=2).border = borde_fino
    row += 2

    # Secciones D.1 y D.2 en horizontal
    row = write_horizontal_section(ws, row, data["seccion_d1"]["titulo"], data["seccion_d1"]["rows"])
    row = write_horizontal_section(ws, row, data["seccion_d2"]["titulo"], data["seccion_d2"]["rows"])

    # --- Sección D.3 en vertical ---
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell_titulo_d3 = ws.cell(row=row, column=1, value=data["seccion_d3"]["titulo"])
    cell_titulo_d3.font = Font(bold=True, size=12)
    cell_titulo_d3.alignment = Alignment(horizontal="center")
    cell_titulo_d3.border = borde_fino
    row += 1

    # Encabezados
    ws.cell(row=row, column=1, value="Indicador").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Valor").font = Font(bold=True)
    ws.cell(row=row, column=1).border = borde_fino
    ws.cell(row=row, column=2).border = borde_fino
    row += 1

    # Filas con colores en valores
    for indicador, valor in data["seccion_d3"]["rows"]:
        cell_indicador = ws.cell(row=row, column=1, value=indicador)
        cell_valor = ws.cell(row=row, column=2, value=valor)
        cell_indicador.border = borde_fino
        cell_valor.border = borde_fino
        cell_valor.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        cell_valor.font = Font(bold=True)
        cell_valor.alignment = Alignment(horizontal="center")
        row += 1

    # Nota de lactancia
    #ws.cell(row=row, column=1, value="Lactancia materna en los primeros 60 minutos de vida (RN ≥ 2500 grs)")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1).font = Font(italic=True, size=10)
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
    ws.cell(row=row, column=1).border = borde_fino

    # Pie de página institucional
    row += 2
    ws.cell(row=row, column=1, value="Fuente: Sistema Neonatal — Hospital Herminda Martin")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1).font = Font(size=9, italic=True, color="808080")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")

    # Respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="REM_A24_{titulo_mes}_{anio}.xlsx"'
    

    wb.save(response)
    return response




def exportar_rem_a24_pdf(request):
    # --- Capturar filtros desde POST o GET ---
    mes = request.POST.get("mes") or request.GET.get("mes")
    anio = request.POST.get("anio") or request.GET.get("anio")
    inicio = request.POST.get("inicio") or request.GET.get("inicio")
    fin = request.POST.get("fin") or request.GET.get("fin")

    mes = int(mes) if mes else None
    anio = int(anio) if anio else None
    inicio = date.fromisoformat(inicio) if inicio else None
    fin = date.fromisoformat(fin) if fin else None

    # --- Obtener datos del reporte ---
    data = get_reporte_rem24_completo(mes=mes, anio=anio, inicio=inicio, fin=fin)

    # --- Renderizar template HTML ---
    html = render_to_string("reportes/exportable_rem_a24.html", {
        "titulo": data["titulo"],
        "periodo": data["periodo"],
        "seccion_d1": data["seccion_d1"],
        "seccion_d2": data["seccion_d2"],
        "seccion_d3": data["seccion_d3"],
    })

    # --- Convertir a PDF ---
    result = io.BytesIO()
    pisa.CreatePDF(html, dest=result)

    # --- Respuesta HTTP ---
    response = HttpResponse(result.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="REM_A24_{data["periodo"]}.pdf"'
    return response

def rem_24(request):
    mes_str = request.GET.get("mes")
    anio_str = request.GET.get("anio")

    now = datetime.now()
    mes = int(mes_str) if mes_str and mes_str.isdigit() else now.month
    anio = int(anio_str) if anio_str and anio_str.isdigit() else now.year

    periodo_legible = f"{nombre_mes(mes)} {anio}"

    data = get_reporte_rem24_completo(mes=mes, anio=anio)

    lista_anios = range(now.year - 5, now.year + 1)

    return render(request, "reportes/rem_24.html", {
        "titulo": data["titulo"],
        "periodo": periodo_legible,
        "seccion_d1": data["seccion_d1"],
        "seccion_d2": data["seccion_d2"],
        "seccion_d3": data["seccion_d3"],
        "lista_anios": lista_anios,
        "mes_seleccionado": mes,
        "anio_seleccionado": anio,
    })